"""
review_traces.py
────────────────
Interactive trace reviewer for agent-failure-detection dataset.

Usage:
    python review_traces.py                        # full review with Claude API annotations
    python review_traces.py --no-api               # label + confidence only, no API calls
    python review_traces.py --folder DISPUTED      # review one class folder only
    python review_traces.py --stats                # show progress stats and exit

Writes:
    - label_human, confidence_human, reviewed, key_evidence,
      failure_pattern, eval_notes  → back into each JSON file
    - One row per trace             → trace_annotation_log.xlsx
"""

import os
import json
import argparse
import anthropic
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# ── Config ───────────────────────────────────────────────────────────────────

DATA_DIR   = Path("data/labelled")
EXCEL_PATH = Path("trace_annotation_log.xlsx")

VALID_LABELS = [
    "SUCCESS", "HALLUCINATION", "GOAL_DRIFT",
    "TOOL_MISUSE", "LOOP", "UNSAFE_EXECUTION", "DISPUTED"
]

LABEL_COLOURS = {
    "SUCCESS":          "C6EFCE",
    "HALLUCINATION":    "FFEB9C",
    "GOAL_DRIFT":       "FFC7CE",
    "TOOL_MISUSE":      "FCE4D6",
    "LOOP":             "DDEBF7",
    "UNSAFE_EXECUTION": "E2EFDA",
    "DISPUTED":         "D9D9D9",
}

CONF_MAP = {"H": "HIGH", "M": "MEDIUM", "L": "LOW"}

# ── Helpers ──────────────────────────────────────────────────────────────────

def collect_traces(folder_filter=None):
    """Return list of (path, data) for all unreviewed traces."""
    traces = []
    if not DATA_DIR.exists():
        print(f"[ERROR] {DATA_DIR} not found. Run from project root.")
        raise SystemExit(1)

    folders = sorted(DATA_DIR.iterdir())
    for folder in folders:
        if not folder.is_dir():
            continue
        if folder_filter and folder.name != folder_filter:
            continue
        for fpath in sorted(folder.glob("*.json")):
            with open(fpath, encoding="utf-8") as f:
                try:
                    data = json.load(f)
                except json.JSONDecodeError:
                    print(f"[WARN] Skipping malformed JSON: {fpath}")
                    continue
            if not data.get("reviewed", False):
                traces.append((fpath, data))
    return traces


def show_stats():
    """Print per-class reviewed/total counts."""
    print("\n── Dataset Review Progress ─────────────────────────────")
    total_done = total_all = 0
    for folder in sorted(DATA_DIR.iterdir()):
        if not folder.is_dir():
            continue
        files = list(folder.glob("*.json"))
        done = sum(
            1 for f in files
            if json.load(open(f, encoding="utf-8")).get("reviewed", False)
        )
        total_done += done
        total_all  += len(files)
        bar = "█" * done + "░" * (len(files) - done)
        print(f"  {folder.name:<20} {done:>3}/{len(files):<3}  {bar}")
    print(f"\n  TOTAL: {total_done}/{total_all} reviewed")
    print("────────────────────────────────────────────────────────\n")


def wrap(text, width=100, indent="    "):
    """Word-wrap text at width, indenting continuation lines."""
    import textwrap
    text = str(text).replace("\n", " ").strip()
    lines = textwrap.wrap(text, width=width)
    if not lines:
        return ""
    return ("\n" + indent).join(lines)


def render_trace(data, folder_name, index, total):
    """Pretty-print a trace to terminal with full word-wrapped content."""
    sep  = "═" * 90
    thin = "─" * 90
    print(f"\n{sep}")
    print(f"  [{index}/{total}]  Trace ID : {data.get('trace_id','?')}")
    print(f"  Folder   : {folder_name}")
    print(f"  Task     : {wrap(data.get('task','?'), width=85)}")
    print(f"  Steps    : {data.get('step_count', len(data.get('steps',[])))}")
    lc = data.get('label_claude','?')
    lg = data.get('label_gpt4o','?')
    cc = data.get('confidence_claude','?')
    cg = data.get('confidence_gpt4o','?')
    print(f"  Auto     : Claude={lc} ({cc})  |  GPT-4o={lg} ({cg})")
    print(thin)
    for i, step in enumerate(data.get("steps", []), 1):
        thought = step.get("thought", "").strip()
        if thought and thought != "[no explicit reasoning]":
            print(f"\n  Step {i} ── THOUGHT")
            print(f"    {wrap(thought, width=85)}")
        print(f"\n  Step {i} ── ACTION  : {step.get('action','?')}")
        print(f"    Input : {wrap(str(step.get('action_input','')), width=85)}")
        obs = str(step.get("observation",""))
        print(f"    Obs   : {wrap(obs, width=85)}")
    print(f"\n{thin}")
    fa = str(data.get("final_answer",""))
    print(f"\n  FINAL ANSWER")
    print(f"    {wrap(fa, width=85)}")
    print(f"\n{sep}")


def get_label():
    """Prompt user for a valid label."""
    shortcuts = {
        "s": "SUCCESS", "h": "HALLUCINATION", "g": "GOAL_DRIFT",
        "t": "TOOL_MISUSE", "l": "LOOP", "u": "UNSAFE_EXECUTION", "d": "DISPUTED"
    }
    print("\n  Labels: [S]uccess  [H]allucination  [G]oal_drift")
    print("          [T]ool_misuse  [L]oop  [U]nsafe  [D]isputed  [skip]")
    while True:
        raw = input("  Your label: ").strip().lower()
        if raw == "skip":
            return None
        if raw in shortcuts:
            return shortcuts[raw]
        up = raw.upper()
        if up in VALID_LABELS:
            return up
        print("  Invalid. Try again.")


def get_confidence():
    """Prompt user for confidence level."""
    while True:
        raw = input("  Confidence [H/M/L]: ").strip().upper()
        if raw in CONF_MAP:
            return CONF_MAP[raw]
        print("  Enter H, M, or L.")


def call_claude_annotation(data, human_label, human_confidence):
    """Call Claude API to generate key evidence, failure pattern, eval notes."""
    client = anthropic.Anthropic()

    steps_text = ""
    for i, step in enumerate(data.get("steps", []), 1):
        steps_text += f"Step {i}: {step.get('action')} | Input: {step.get('action_input')} | Obs: {str(step.get('observation',''))[:300]}\n"

    prompt = f"""You are annotating an LLM agent execution trace for a research dataset on agent failure classification.

Task: {data.get('task')}

Steps:
{steps_text}

Final Answer: {data.get('final_answer')}

The human reviewer has labelled this trace as: {human_label} (confidence: {human_confidence})

Based on the trace content and the assigned label, provide exactly three things:

KEY_EVIDENCE: One or two specific sentences pointing to the exact step(s) or content that justifies the label. Be precise — reference step numbers and actual content.

FAILURE_PATTERN: A short phrase (5-10 words) naming the specific failure pattern observed. Examples: "Semantic loop with no state change", "Fabricated statistic not in tool output", "Task topic shifted at step 3".

EVAL_NOTES: One or two sentences noting anything interesting about this trace for classifier evaluation — e.g. edge cases, ambiguity, what makes it hard or easy to classify, or what distinguishes it from similar failure types.

Respond in exactly this format:
KEY_EVIDENCE: ...
FAILURE_PATTERN: ...
EVAL_NOTES: ..."""

    message = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=400,
        messages=[{"role": "user", "content": prompt}]
    )

    text = message.content[0].text.strip()
    result = {"key_evidence": "", "failure_pattern": "", "eval_notes": ""}

    for line in text.split("\n"):
        if line.startswith("KEY_EVIDENCE:"):
            result["key_evidence"] = line.replace("KEY_EVIDENCE:", "").strip()
        elif line.startswith("FAILURE_PATTERN:"):
            result["failure_pattern"] = line.replace("FAILURE_PATTERN:", "").strip()
        elif line.startswith("EVAL_NOTES:"):
            result["eval_notes"] = line.replace("EVAL_NOTES:", "").strip()

    return result


def save_json(fpath, data, human_label, human_confidence, annotations):
    """Write human review fields back into the JSON file."""
    data["label_human"]       = human_label
    data["confidence_human"]  = human_confidence
    data["final_label"]       = human_label  # update final label to human verdict
    data["reviewed"]          = True
    data["reviewed_at"]       = datetime.utcnow().isoformat()
    data["key_evidence"]      = annotations.get("key_evidence", "")
    data["failure_pattern"]   = annotations.get("failure_pattern", "")
    data["eval_notes"]        = annotations.get("eval_notes", "")
    with open(fpath, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def ensure_excel():
    """Create the Excel file with headers if it does not exist."""
    from openpyxl.utils import get_column_letter
    if EXCEL_PATH.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Trace Annotations"
    headers = [
        "Trace ID", "Original Label", "Verified Label",
        "Confidence", "Key Evidence", "Failure Pattern", "Eval Notes", "Trace Content"
    ]
    header_fill = PatternFill("solid", start_color="1F3864")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    for col, w in enumerate([12, 18, 18, 12, 40, 30, 40, 80], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    wb.save(EXCEL_PATH)
    print(f"  ✓ Created {EXCEL_PATH}")


def append_to_excel(data, folder_name, human_label, human_confidence, annotations):
    """Append one row to the Excel annotation log."""
    ensure_excel()
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Trace Annotations"]
    next_row = ws.max_row + 1

    original_label = folder_name  # always the source folder name

    # Build full trace for the cell — no truncation
    steps_summary = ""
    for i, step in enumerate(data.get("steps", []), 1):
        thought = step.get("thought", "").strip()
        if thought and thought != "[no explicit reasoning]":
            steps_summary += f"[{i}] THOUGHT: {thought}\n"
        steps_summary += f"[{i}] ACTION: {step.get('action')}\n"
        steps_summary += f"[{i}] INPUT: {step.get('action_input','')}\n"
        steps_summary += f"[{i}] OBS: {str(step.get('observation',''))}\n\n"
    trace_content = f"TASK: {data.get('task')}\n\n{steps_summary}FINAL: {data.get('final_answer','')}"

    row_data = [
        data.get("trace_id", "")[:8],
        original_label,
        human_label,
        human_confidence,
        annotations.get("key_evidence", ""),
        annotations.get("failure_pattern", ""),
        annotations.get("eval_notes", ""),
        trace_content,
    ]

    left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center     = Alignment(horizontal="center", vertical="center")

    for c, val in enumerate(row_data, 1):
        cell = ws.cell(row=next_row, column=c, value=val)
        cell.font = Font(name="Arial", size=9)
        cell.alignment = left_wrap if c in (5, 6, 7, 8) else center

    # Colour the verified label cell
    colour = LABEL_COLOURS.get(human_label, "FFFFFF")
    ws.cell(row=next_row, column=3).fill = PatternFill("solid", start_color=colour)

    wb.save(EXCEL_PATH)


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--no-api",  action="store_true", help="Skip Claude API annotation")
    parser.add_argument("--folder",  type=str, default=None, help="Review one folder only")
    parser.add_argument("--stats",   action="store_true", help="Show progress and exit")
    args = parser.parse_args()

    if args.stats:
        show_stats()
        return

    # Check API key
    import os
    if not os.getenv("ANTHROPIC_API_KEY"):
        print("\n[ERROR] ANTHROPIC_API_KEY not found in environment or .env file.")
        print("  Either set it in your .env file or run with --no-api flag.")
        raise SystemExit(1)
    else:
        key = os.getenv("ANTHROPIC_API_KEY")
        print(f"  ✓ API key loaded: sk-...{key[-6:]}")

    traces = collect_traces(folder_filter=args.folder)

    if not traces:
        print("\n✓ All traces reviewed!" if not args.folder else f"\n✓ All traces in {args.folder} reviewed!")
        show_stats()
        return

    print(f"\n{'═'*70}")
    print(f"  TRACE REVIEW SESSION")
    print(f"  {len(traces)} unreviewed traces")
    if args.no_api:
        print("  Mode: label + confidence only (--no-api)")
    else:
        print("  Mode: full annotation via Claude API")
    print(f"  Excel: {EXCEL_PATH}")
    print(f"{'═'*70}")
    print("  Type 'skip' to skip a trace, Ctrl+C to stop session.\n")

    reviewed_count = 0

    for i, (fpath, data) in enumerate(traces, 1):
        folder_name = fpath.parent.name

        render_trace(data, folder_name, i, len(traces))

        human_label = get_label()
        if human_label is None:
            print("  ↷ Skipped.\n")
            continue

        human_confidence = get_confidence()

        if args.no_api:
            annotations = {"key_evidence": "", "failure_pattern": "", "eval_notes": ""}
        else:
            print("  ⟳ Calling Claude API for annotations...")
            try:
                annotations = call_claude_annotation(data, human_label, human_confidence)
                print(f"  KEY EVIDENCE   : {annotations['key_evidence']}")
                print(f"  FAILURE PATTERN: {annotations['failure_pattern']}")
                print(f"  EVAL NOTES     : {annotations['eval_notes']}")
            except Exception as e:
                import traceback
                print(f"\n  [ERROR] API call failed:")
                traceback.print_exc()
                print()
                annotations = {"key_evidence": "", "failure_pattern": "", "eval_notes": ""}

        save_json(fpath, data, human_label, human_confidence, annotations)
        append_to_excel(data, folder_name, human_label, human_confidence, annotations)

        reviewed_count += 1
        print(f"\n  ✓ Saved → {fpath.name}\n")

    print(f"\n{'═'*70}")
    print(f"  Session complete. {reviewed_count} traces reviewed.")
    show_stats()


if __name__ == "__main__":
    main()